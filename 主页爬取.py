import time
from bs4 import BeautifulSoup
import re
import requests
import openpyxl


html = 0
'''
取得所要爬取页面的HTML源码
'''
def askURL(url):
    global html
    try:
        response = requests.get(url)
        print(response.status_code)
        html = response.text
    except Exception as e:
        print(e)
    return html


'''
使用正则表达式截取相关内容
此部分与网页结构密切相关，若相关部分结构改变，规则亦需要改变
'''
def getData(baseurl):
    global html
    findPage = re.compile(r'<a href="(.+\n*)" target=".+\n*">')  # 找到主页中各个副页文章链接
    findTitle = re.compile(r'<a href=".+\n*" target=".+\n*">(.+\n*)</a>')  # 找到游记题名
    findName = re.compile(r'<a href=".+\n*" title=".+\n*">(.+\n*)</a>')  # 找到文章作者昵称
    findJourney = re.compile(r'</a>(.+)公里')  # 找到作者行程里数
    findStartTime = re.compile(r'于(.+)出发')  # 找到作者旅游出发时间
    findSpentTime = re.compile(r'历时(.+)\n\s+</dd>')  # 找到作者旅游旅游时长
    findImgSrc = re.compile(r'<img.*src="(.*)"\n*\s*title')  # 找到展示图片连接，列表
    datalist = []  # 每个游客信息为一项
    pagelist = []  # 副页文章链接列表

    for i in range(1, 10):  # 以'河北白石山'为关键词搜索，检索结构9页
        url = baseurl.replace('page_num', str(i))
        print(url)
        html = askURL(url)
        soup = BeautifulSoup(html, "html.parser")
        j = 1  # 主页中游客文章前面有一个与游客文章相同的盒子，避开
        for item in soup.find_all('table', width="100%", border="0",
                                  cellspacing="0", cellpadding="0"):  # 找到每一个文章项
            if j == 1:
                j += 1
                continue
            data = []
            item = str(item)  # 转换成字符串
            # print(item)
            page = 'http://www.foooooot.com' + re.findall(findPage, item)[0]  # 取得每个主页链接
            pagelist.append(page)
            title = re.findall(findTitle, item)[0]
            data.append(title)
            name = re.findall(findName, item)[0]
            data.append(name)
            journey = re.findall(findJourney, item)[0]
            data.append(journey)
            start_time = re.findall(findStartTime, item)[0]
            data.append(start_time)
            spent_time = re.findall(findSpentTime, item)[0]
            data.append(spent_time)
            imgs_src = re.findall(findImgSrc, item)
            imgs_src = '\n'.join(imgs_src)  # 将每个照片链接换行显示
            data.append(imgs_src)
            datalist.append(data)
    time.sleep(5)
    return datalist, pagelist


'''
保存到Excel表格
'''
def saveData(datalist, savepath):
    book = openpyxl.Workbook()
    sheet = book.create_sheet("河北白石山六只脚游记主页Top247")
    col = ('游记题名', '文章作者昵称', '作者行程/公里', '作者旅游出发时间', '作者旅游旅游时长', '展示图片')
    sheet.append(col)

    for i in range(0, 247):
        data = datalist[i]
        for j in range(0, 6):  # 从0开始计算，注意去掉列头
            sheet.cell(row=(i + 2), column=(j + 1), value=data[j])
    book.save(savepath)


def main():
    print("开始爬取......")
    baseurl = 'http://www.foooooot.com/search/trip/all/1/with_pics/default/descent/?page=page_num&keyword=%E6%B2%B3%E5%8C%97%E7%99%BD%E7%9F%B3%E5%B1%B1'
    datalist, pagelist = getData(baseurl)
    with open('data/pagelist.txt', 'w') as pages:
            pages.write('\n'.join(pagelist))
    savapath = u'data/河北白石山六只脚游记主页Top247.xlsx'
    saveData(datalist, savapath)


if __name__ == '__main__':
    main()
    print("结束爬取......data/河北白石山六只脚游记主页Top247.xlsx")