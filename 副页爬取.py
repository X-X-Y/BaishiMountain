import time
from bs4 import BeautifulSoup
import re
import requests
import openpyxl


html = 0
'''
取得所要爬取页面的HTML源码
'''
def askURL(url):
    global html
    try:
        response = requests.get(url)
        print(response.status_code)
        html = response.text
    except Exception as e:
        print(e)
    return html


'''
使用正则表达式截取相关内容
此部分与网页结构密切相关，若相关部分结构改变，规则亦需要改变
'''
def getData():
    global html
    data1 = []
    data2 = []
    data3 = []
    findLink = re.compile(r'<h1 class="title">\n*(.+\n*)</h1>')  # 找到游记题名
    findName_Way = re.compile(r'<a href=".+\n*" title=".+\n*">(.+\n*)</a>')  # 找到文章作者昵称和旅游形式
    findJourney = re.compile(r'全程(.+)公里')  # 找到作者行程里数

    findRise = re.compile(r'累计上升</strong>：(.+)米，<strong>')  # 找到作者累计上升
    findDescend = re.compile(r'累计下降：</strong>(.+)米')  # 找到作者累计下降
    findLow= re.compile(r'<span class="low">(.+)</span>米，<strong>')  # 找到作者海拔最低
    findHeight = re.compile(r'<span class="height">(.+)</span>')  # 找到作者海拔最高
    findSpeed = re.compile(r'</strong>(.+)公里每小时')  # 找到作者最高速度

    findStartTime = re.compile(r'于(.+)出发')  # 找到作者旅游出发时间
    findSpentTime = re.compile(r'历时(.+)\n\s+</dd>')  # 找到作者旅游旅游时长

    findImgSrc = re.compile(r'<span class="down_img" download="sixfoot.jpg" href="(.+)">')  # 找到展示图片连接
    findLongitude = re.compile(r'<span class="lat_lng" title="经纬度">(.+)</span>')  # 找到展示图片经纬度
    datalist = []


    with open('data/pagelist1.txt') as q:  # 这里就用到主页爬取保存下来的文章连接了
        lines = q.readlines()
        for i in range(len(lines)):
            lines[i] = lines[i].replace('\n', '')
        for line in lines:
            html = askURL(line)
            soup = BeautifulSoup(html, "html.parser")
            for item1 in soup.find_all('div', class_="trip_box trip_box_title"):  # 找到每一个文章标题项
                data1 = []
                item1 = str(item1)
                # print(item1)
                # 这里少数也会报错，let me see see
                try:
                    link = re.findall(findLink, item1)[0]
                except:
                    print(item1)
                data1.append(link)
            for item2 in soup.find_all('dl', class_="trip_box_right"):  # 找到每一个文章简介项
                data2 = []
                item2 = str(item2)
                # print(item2)
                name = re.findall(findName_Way, item2)[0]
                data2.append(name)
                way = re.findall(findName_Way, item2)[1]
                data2.append(way)
                journey = re.findall(findJourney, item2)[0]
                data2.append(journey)
                rise = re.findall(findRise, item2)[0]
                data2.append(rise)
                descend = re.findall(findDescend, item2)[0]
                data2.append(descend)
                low = re.findall(findLow, item2)[0]
                data2.append(low)
                height = re.findall(findHeight, item2)[0]
                data2.append(height)
                speed = re.findall(findSpeed, item2)[0]
                data2.append(speed)
                start_time = re.findall(findStartTime, item2)[0]
                data2.append(start_time)

                findMonth = re.compile('-(\d+)-')  # 格式为' 2017-12-09 08:50 '
                season = re.findall(findMonth, start_time)
                if int(season[0]) >= 3 and int(season[0]) <= 5:
                    season = '春季'
                elif int(season[0]) >= 6 and int(season[0]) <= 8:
                    season = '夏季'
                elif int(season[0]) >= 9 and int(season[0]) <= 11:
                    season = '秋季'
                else:
                    season = '冬季'
                data2.append(season)

                spent_time = re.findall(findSpentTime, item2)[0]
                data2.append(spent_time)
            for item3 in soup.find_all('div', class_="key"):  # 找到每一个文章照片项
                item3 = str(item3)
                # print(item3)
                imgs_src = re.findall(findImgSrc, item3)
                longitude = re.findall(findLongitude, item3)
                # 这里少数会有报错，因为那些文章没有照片或者没有经纬度坐标，嗯，坑呐
                try:
                    imgs_info = imgs_src[0] + '    ' + longitude[0]
                except:
                    imgs_info = '照片或经纬度信息缺失'  # 放弃该篇文章照片信息
                data3.append(imgs_info)
            data3[0] = '\n'.join(data3)
            data = data1 + data2 + data3
            datalist.append(data)
    time.sleep(5)
    return datalist


'''
保存到Excel表格
'''
def saveData(datalist, savepath):
    book = openpyxl.Workbook()
    sheet = book.create_sheet("河北白石山六只脚游记副页Top")
    col = ('游记题名', '文章作者昵称', '旅游形式', '作者行程(公里)',
           '累计上升(米)', '累计下降(米)', '海拔最低(米)', '海拔最高(米)', '最高速度(公里/h)',
           '作者旅游出发时间', '旅游季节', '作者旅游旅游时长', '展示图片')
    sheet.append(col)

    for i in range(0, 247):
        data = datalist[i]
        for j in range(0, 13):  # 从0开始计算，注意去掉列头
            sheet.cell(row=(i + 2), column=(j + 1), value=data[j])
    book.save(savepath)  # 保存


def main():
    print("开始爬取......")
    datalist=getData()
    savapath=u'data/河北白石山六只脚游记副页Top247.xlsx'
    saveData(datalist, savapath)


if __name__ == '__main__':
    main()
    print("结束爬取......data/河北白石山六只脚游记副页Top247.xlsx")